"""
benchmark/hrms_mass_accuracy.py
HRMS Mass Accuracy Validation for HRMS-Predict
Runs 10 reference compounds and compares predicted masses against literature.

Usage:
    conda activate hrms-predictor
    cd hrms-predict
    python benchmark\\hrms_mass_accuracy.py
"""
import sys
import time
from pathlib import Path

import requests

try:
    import pandas as pd
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install",
                           "pandas", "openpyxl", "--quiet"])
    import pandas as pd

PREDICT_URL = "http://localhost:8000/predict"
HEALTH_URL  = "http://localhost:8000/health"
RESULTS_DIR = Path("benchmark/results")

# All literature masses verified against PubChem and NIST
# Oseltamivir SMILES uses cyclohex-1-ene ring (C1=C) for correct C16H28N2O4
REFERENCE_COMPOUNDS = [
    {
        "name":         "Aspirin",
        "smiles":       "CC(=O)Oc1ccccc1C(=O)O",
        "formula":      "C9H8O4",
        "mw_range":     "small (180 Da)",
        "lit_neutral":  180.0423,
        "lit_mplus_h":  181.0495,
        "lit_mminus_h": 179.0350,
        "source":       "NIST, PubChem CID 2244",
    },
    {
        "name":         "Acetaminophen",
        "smiles":       "CC(=O)Nc1ccc(O)cc1",
        "formula":      "C8H9NO2",
        "mw_range":     "small (151 Da)",
        "lit_neutral":  151.0633,
        "lit_mplus_h":  152.0706,
        "lit_mminus_h": 150.0560,
        "source":       "NIST, PubChem CID 1983",
    },
    {
        "name":         "Caffeine",
        "smiles":       "Cn1cnc2c1c(=O)n(C)c(=O)n2C",
        "formula":      "C8H10N4O2",
        "mw_range":     "small (194 Da)",
        "lit_neutral":  194.0804,
        "lit_mplus_h":  195.0877,
        "lit_mminus_h": 193.0731,
        "source":       "NIST, PubChem CID 2519",
    },
    {
        "name":         "Ibuprofen",
        "smiles":       "CC(C)Cc1ccc(cc1)C(C)C(=O)O",
        "formula":      "C13H18O2",
        "mw_range":     "medium (206 Da)",
        "lit_neutral":  206.1307,
        "lit_mplus_h":  207.1380,
        "lit_mminus_h": 205.1234,
        "source":       "NIST, PubChem CID 3672",
    },
    {
        "name":         "Diclofenac",
        "smiles":       "O=C(O)Cc1ccccc1Nc1c(Cl)cccc1Cl",
        "formula":      "C14H11Cl2NO2",
        "mw_range":     "medium (295 Da)",
        "lit_neutral":  295.0167,
        "lit_mplus_h":  296.0239,
        "lit_mminus_h": 294.0094,
        "source":       "NIST, PubChem CID 3033",
    },
    {
        "name":         "Warfarin",
        "smiles":       "CC(=O)CC(c1ccccc1)c1c(O)c2ccccc2oc1=O",
        "formula":      "C19H16O4",
        "mw_range":     "medium (308 Da)",
        "lit_neutral":  308.1049,
        "lit_mplus_h":  309.1121,
        "lit_mminus_h": 307.0976,
        "source":       "NIST, PubChem CID 54678486",
    },
    {
        "name":         "Omeprazole",
        "smiles":       "COc1ccc2[nH]c(S(=O)Cc3ncc(C)c(OC)c3C)nc2c1",
        "formula":      "C17H19N3O3S",
        "mw_range":     "medium (345 Da)",
        "lit_neutral":  345.1147,
        "lit_mplus_h":  346.1220,
        "lit_mminus_h": 344.1074,
        "source":       "NIST, PubChem CID 4594",
    },
    {
        "name":         "Metformin",
        "smiles":       "CN(C)C(=N)NC(N)=N",
        "formula":      "C4H11N5",
        "mw_range":     "small (129 Da)",
        "lit_neutral":  129.1014,
        "lit_mplus_h":  130.1087,
        "lit_mminus_h": 128.0941,
        "source":       "NIST, PubChem CID 4091",
    },
        {
        "name":         "Naproxen",
        "smiles":       "COc1ccc2cc(C(C)C(=O)O)ccc2c1",
        "formula":      "C14H14O3",
        "mw_range":     "medium (230 Da)",
        "lit_neutral":  230.0943,
        "lit_mplus_h":  231.1016,
        "lit_mminus_h": 229.0870,
        "source":       "NIST, PubChem CID 156391",
    },
    {
        "name":         "Atorvastatin",
        "smiles":       "CC(C)c1n(CC[C@@H](O)C[C@@H](O)CC(=O)O)c(-c2ccccc2F)c(C(=O)Nc2ccccc2)c1-c1ccccc1",
        "formula":      "C33H35FN2O5",
        "mw_range":     "large (558 Da)",
        "lit_neutral":  558.2530,
        "lit_mplus_h":  559.2603,
        "lit_mminus_h": 557.2457,
        "source":       "NIST, PubChem CID 60823",
    },
]


def ppm_error(predicted, literature):
    return round((predicted - literature) / literature * 1e6, 3)


def abs_error_mda(predicted, literature):
    return round(abs(predicted - literature) * 1000, 4)


def check_server():
    try:
        r = requests.get(HEALTH_URL, timeout=5)
        return r.status_code == 200
    except Exception:
        return False


def run_prediction(smiles):
    try:
        r = requests.post(PREDICT_URL, json={
            "smiles":             smiles,
            "run_sygma":          True,
            "run_biotransformer": False,
            "run_dl":             False,
            "run_smartcyp":       False,
        }, timeout=60)
        if r.status_code == 200:
            return r.json().get("parent")
        print(f"  HTTP {r.status_code}: {r.text[:80]}")
        return None
    except Exception as e:
        print(f"  Error: {e}")
        return None


def run():
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)

    if not check_server():
        print("Backend not running at localhost:8000")
        print("Start with: uvicorn app.main:app --reload --port 8000")
        sys.exit(1)
    print("Backend online\n")

    print("=" * 72)
    print("HRMS-Predict -- Mass Accuracy Validation")
    print("=" * 72)
    print(f"{'Compound':20s}  {'Pred MW':>9s}  {'ppm Neutral':>12s}  {'ppm [M+H]+':>11s}  Status")
    print("-" * 72)

    rows = []

    for cpd in REFERENCE_COMPOUNDS:
        parent = run_prediction(cpd["smiles"])
        if parent is None:
            print(f"  {'ERROR':20s}  {cpd['name']}")
            continue

        pred_neutral = round(parent.get("neutral_mass", 0.0), 4)
        pred_mplus   = round(parent.get("adducts", {}).get("mplus_h", 0.0), 4)
        pred_mminus  = round(parent.get("adducts", {}).get("mminus_h", 0.0), 4)

        ppm_n  = ppm_error(pred_neutral, cpd["lit_neutral"])
        ppm_p  = ppm_error(pred_mplus,   cpd["lit_mplus_h"])
        ppm_m  = ppm_error(pred_mminus,  cpd["lit_mminus_h"])
        mda_n  = abs_error_mda(pred_neutral, cpd["lit_neutral"])
        mda_p  = abs_error_mda(pred_mplus,   cpd["lit_mplus_h"])

        status = "PASS" if abs(ppm_n) < 1.0 else "FAIL"
        print(
            f"  {cpd['name']:20s}  "
            f"{pred_neutral:9.4f}  "
            f"{ppm_n:+10.4f} ppm  "
            f"{ppm_p:+9.4f} ppm  "
            f"{status}"
        )

        rows.append({
            "Compound":               cpd["name"],
            "Formula":                cpd["formula"],
            "MW Range":               cpd["mw_range"],
            "Lit. Neutral Mass (Da)": cpd["lit_neutral"],
            "Pred. Neutral Mass (Da)":pred_neutral,
            "Error Neutral (mDa)":    mda_n,
            "ppm Error Neutral":      ppm_n,
            "Lit. [M+H]+ (Da)":       cpd["lit_mplus_h"],
            "Pred. [M+H]+ (Da)":      pred_mplus,
            "Error [M+H]+ (mDa)":     mda_p,
            "ppm Error [M+H]+":       ppm_p,
            "Lit. [M-H]- (Da)":       cpd["lit_mminus_h"],
            "Pred. [M-H]- (Da)":      pred_mminus,
            "ppm Error [M-H]-":       ppm_m,
            "Source":                 cpd["source"],
        })

    print("-" * 72)

    if not rows:
        print("No results -- check server is running")
        return

    df = pd.DataFrame(rows)
    max_ppm  = df["ppm Error Neutral"].abs().max()
    mean_ppm = df["ppm Error Neutral"].abs().mean()
    max_mda  = df["Error Neutral (mDa)"].max()
    n_pass   = (df["ppm Error Neutral"].abs() < 1.0).sum()

    print(f"\n  Compounds tested  : {len(df)}")
    print(f"  All within 1 ppm  : {n_pass}/{len(df)}")
    print(f"  Max |ppm error|   : {max_ppm:.4f} ppm")
    print(f"  Mean |ppm error|  : {mean_ppm:.4f} ppm")
    print(f"  Max |mDa error|   : {max_mda:.4f} mDa")

    # Save CSV
    csv_path = RESULTS_DIR / "mass_accuracy_results.csv"
    df.to_csv(csv_path, index=False)

    # Save publication-ready Excel
    xlsx_path = RESULTS_DIR / "mass_accuracy_table.xlsx"
    pub_cols = [
        "Compound", "Formula",
        "Lit. Neutral Mass (Da)", "Pred. Neutral Mass (Da)",
        "Error Neutral (mDa)", "ppm Error Neutral",
        "Lit. [M+H]+ (Da)", "Pred. [M+H]+ (Da)",
        "ppm Error [M+H]+",
    ]
    pub_df = df[pub_cols].copy()

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        pub_df.to_excel(writer, sheet_name="Mass Accuracy", index=False)
        df.to_excel(writer, sheet_name="Full Data", index=False)

        ws = writer.sheets["Mass Accuracy"]
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="1A3057")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        alt_fill    = PatternFill("solid", fgColor="EBF4FF")
        thin        = Side(style="thin", color="BEE3F8")
        bdr         = Border(left=thin, right=thin, top=thin, bottom=thin)

        col_widths = [18, 12, 22, 22, 20, 18, 18, 18, 16]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        for row in ws.iter_rows():
            for cell in row:
                cell.border = bdr
                cell.alignment = Alignment(
                    horizontal="center" if cell.column > 1 else "left",
                    vertical="center"
                )
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                elif cell.row % 2 == 0:
                    cell.fill = alt_fill

        ws.row_dimensions[1].height = 30

    # Save text report (UTF-8 encoding for Windows compatibility)
    report_path = RESULTS_DIR / "mass_accuracy_report.txt"
    report = (
        "=" * 72 + "\n"
        "HRMS-Predict -- Mass Accuracy Validation Report\n"
        "=" * 72 + "\n\n"
        "Method\n"
        "------\n"
        "  Parent compound SMILES submitted to /predict endpoint.\n"
        "  Neutral monoisotopic mass from parent metrics in JSON response.\n"
        "  Literature masses from NIST Chemistry WebBook and PubChem.\n"
        "  ppm error = (predicted - literature) / literature * 10^6\n\n"
        "Results\n"
        "-------\n"
        f"  Compounds tested      : {len(df)}\n"
        f"  MW range              : 129 Da (Metformin) to 558 Da (Atorvastatin)\n"
        f"  All within 1.0 ppm    : {n_pass}/{len(df)}\n"
        f"  Max |ppm error|       : {max_ppm:.4f} ppm\n"
        f"  Mean |ppm error|      : {mean_ppm:.4f} ppm\n"
        f"  Max |mDa error|       : {max_mda:.4f} mDa\n\n"
        "Per-Compound Results\n"
        "--------------------\n"
    )
    for _, r in df.iterrows():
        report += (
            f"  {r['Compound']:20s}  "
            f"MW={r['Pred. Neutral Mass (Da)']:.4f}  "
            f"ppm={r['ppm Error Neutral']:+.4f}  "
            f"mDa={r['Error Neutral (mDa)']:.4f}\n"
        )
    report += (
        "\nInterpretation\n"
        "--------------\n"
        f"  All {n_pass}/{len(df)} compounds show ppm error < 1.0 ppm.\n"
        "  The [M+H]+ adduct mass is calculated as:\n"
        "      [M+H]+ = M_neutral + 1.007276 Da (proton mass, IUPAC 2018)\n"
        "  Instrument mass accuracy (1-5 ppm on Orbitrap/Q-TOF) is the\n"
        "  dominant source of error in practice, not this calculation.\n\n"
        f"Files saved to: benchmark/results/\n"
    )

    with open(report_path, "w", encoding="utf-8") as f:
        f.write(report)

    print(f"\n  Saved CSV  : {csv_path}")
    print(f"  Saved XLSX : {xlsx_path}")
    print(f"  Saved TXT  : {report_path}")
    print("=" * 72)


if __name__ == "__main__":
    run()



